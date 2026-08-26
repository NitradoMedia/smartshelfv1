"""Parse cash-register export files (CSV / JSON / TXT).

Expected CSV columns (flexible aliases):
  id / transaction_id / bon_id
  timestamp / zeit / datetime / date
  articles / artikel / article_count / anzahl
  total / betrag / amount (optional)
  cashier / kassierer (optional)
  register / kasse (optional)
"""

from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from dateutil import parser as date_parser


@dataclass
class PosTransaction:
    external_id: str
    timestamp: datetime
    article_count: int
    total_amount: Optional[float] = None
    cashier: Optional[str] = None
    register_id: Optional[str] = None
    raw_payload: Optional[str] = None


ID_KEYS = ("id", "transaction_id", "bon_id", "bon", "beleg", "txn_id", "nummer")
TIME_KEYS = ("timestamp", "zeit", "datetime", "date", "datum", "time", "uhrzeit")
ARTICLE_KEYS = (
    "articles",
    "artikel",
    "article_count",
    "anzahl",
    "anzahl_artikel",
    "items",
    "item_count",
    "menge",
)
AMOUNT_KEYS = ("total", "betrag", "amount", "summe", "gesamt")
CASHIER_KEYS = ("cashier", "kassierer", "bediener", "operator")
REGISTER_KEYS = ("register", "kasse", "register_id", "terminal")


def _norm(key: str) -> str:
    return re.sub(r"[^a-z0-9_]", "", key.strip().lower().replace(" ", "_"))


def _pick(row: dict[str, Any], keys: tuple[str, ...]) -> Any:
    normalized = {_norm(k): v for k, v in row.items()}
    for key in keys:
        if key in normalized and normalized[key] not in (None, ""):
            return normalized[key]
    return None


def _parse_ts(value: Any) -> datetime:
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    # German common formats
    for fmt in (
        "%d.%m.%Y %H:%M:%S",
        "%d.%m.%Y %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M:%SZ",
        "%d/%m/%Y %H:%M:%S",
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return date_parser.parse(text, dayfirst=True)


def _parse_int(value: Any) -> int:
    if value is None:
        raise ValueError("article count missing")
    text = str(value).strip().replace(",", ".")
    return int(float(text))


def _parse_float(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    text = str(value).strip().replace("€", "").replace(" ", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def row_to_transaction(row: dict[str, Any], source: str, index: int) -> PosTransaction:
    external = _pick(row, ID_KEYS) or f"{Path(source).stem}-{index}"
    ts = _parse_ts(_pick(row, TIME_KEYS))
    articles = _parse_int(_pick(row, ARTICLE_KEYS))
    return PosTransaction(
        external_id=str(external),
        timestamp=ts,
        article_count=articles,
        total_amount=_parse_float(_pick(row, AMOUNT_KEYS)),
        cashier=(str(_pick(row, CASHIER_KEYS)) if _pick(row, CASHIER_KEYS) else None),
        register_id=(str(_pick(row, REGISTER_KEYS)) if _pick(row, REGISTER_KEYS) else None),
        raw_payload=json.dumps(row, ensure_ascii=False, default=str),
    )


def parse_csv(path: Path) -> list[PosTransaction]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        sample = fh.read(4096)
        fh.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,\t|")
        except csv.Error:
            dialect = csv.excel
            dialect.delimiter = ";" if sample.count(";") > sample.count(",") else ","
        reader = csv.DictReader(fh, dialect=dialect)
        out: list[PosTransaction] = []
        for i, row in enumerate(reader, start=1):
            if not any(row.values()):
                continue
            out.append(row_to_transaction(row, path.name, i))
        return out


def parse_json(path: Path) -> list[PosTransaction]:
    data = json.loads(path.read_text(encoding="utf-8"))
    rows = data if isinstance(data, list) else data.get("transactions") or data.get("bons") or [data]
    return [row_to_transaction(row, path.name, i) for i, row in enumerate(rows, start=1)]


def parse_excel(path: Path) -> list[PosTransaction]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []
    keys = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(header)]
    out: list[PosTransaction] = []
    for i, values in enumerate(rows_iter, start=1):
        if values is None or not any(v is not None and str(v).strip() for v in values):
            continue
        row = {keys[j]: values[j] for j in range(min(len(keys), len(values)))}
        out.append(row_to_transaction(row, path.name, i))
    return out


def parse_pos_file(path: Path) -> list[PosTransaction]:
    suffix = path.suffix.lower()
    if suffix == ".json":
        return parse_json(path)
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return parse_excel(path)
    if suffix in {".csv", ".txt", ".tsv"}:
        return parse_csv(path)
    raise ValueError(f"Unsupported POS file type: {suffix}")
